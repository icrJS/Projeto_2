#pasta venv e activate
#(venv) $ python -m idlelib.idle: novo idle pelo venv
#import os
#os.system("C:\italo\projetos_python\projeto_2\Scripts/activate.bat")

def main():
    from openpyxl import Workbook
    import os

    pasta = os.getcwd()
    try:
        os.chdir(f"{pasta}/planilhas")
    except:
        os.mkdir(f"{pasta}/planilhas")
        os.chdir(f"{pasta}/planilhas")

    wb = Workbook()
    for planilha in ['receitas', 'despesas', 'resultado']:
        wb.create_sheet(planilha)
    print(wb.sheetnames)
    wb.save ('orcamento.xls')
        


if __name__ == "__main__":
    main()
    
    

    




